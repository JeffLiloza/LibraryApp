#######################################################IMPORTS#############################################

from flask import Blueprint, render_template, session, abort, request, redirect, url_for, flash, send_file
import psycopg2  # pip install psycopg2
import psycopg2.extras
import requests
from jsonpath_ng import jsonpath, parse
import urllib.request
import datetime
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os
import re
import telegram_send
from openpyxl import Workbook
###########################################################################################################


#################################Database##########
DB_HOST = "localhost"
DB_NAME = "lib_last"
DB_USER = "superadmin"
DB_PASS = "JeffLiloza"
conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
##################################################


#############################Photo Situation########
UPLOAD_FOLDER = 'static/Photo/'

ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg', 'gif'])


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
####################################################


#################BluePrint########################
adminpn = Blueprint('adminpn', __name__)
###############################################


##################################ROUTES#################################


#############MAIN PAGE###############
@adminpn.route('/adminpn', methods=['POST', 'GET'])
def admin_panel():
    if 'loggedin' in session:
        if session['role'] == 'Admin':
            return render_template('admin/adminpn.html', username=session['username'])
        else:
            return redirect(url_for('userpn.dashboard'))
    else:
        return redirect(url_for('login'))
#####################################


####################STUDENTS STUFF##############################


############STUDENT LIST############
@adminpn.route('/adminpn/students', methods=['POST', 'GET'])
def list_students():
    if 'loggedin' in session:
        if session['role'] == 'Admin':
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            s = "SELECT * FROM students"
            cur.execute(s)  # Execute the SQL
            list_users = cur.fetchall()
            return render_template('/admin/students.html', list_users=list_users, username=session['username'])
        else:
            flash('You dont have an admin control, sorry dude')
            return redirect(url_for('userpn.dashboard'))
    else:
        return redirect(url_for('login'))
####################################


##########STUDENT ADD ##############
@adminpn.route('/adminpn/add_student', methods=['POST', 'GET'])
def add_student():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('adminpn.list_students'))
    else:
        file = request.files['file']
    if file.filename == '':
        flash('No image selected for uploading')
        return redirect(url_for('adminpn.list_students'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        photost = f'Photo/{filename}'
        name = request.form['name']
        grade = request.form['grade']
        card_id = request.form['card_id']
        cur.execute(
            f"INSERT INTO students (name, grade, card_id, photo) VALUES ('{name}','{grade}','{card_id}','{photost}'  )")
        conn.commit()
        flash('Student Added successfully')
        return redirect(url_for('adminpn.list_students'))
    else:
        flash('Allowed image types are - png, jpg, jpeg, gif')
        return redirect(url_for('adminpn.list_students'))

####################################


###########STUDENT EDIT###############
@adminpn.route('/adminpn/student_edit/<id>', methods=['POST', 'GET'])
def get_employee(id):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute(f'SELECT * FROM students WHERE id = {id}')
    data = cur.fetchall()
    cur.close()
    print(data[0])
    return render_template('/admin/student_edit.html', student=data[0])
 ###################################


############STUDENT UPDATE##########
@adminpn.route('/adminpn/student_update/<id>', methods=['POST'])
def update_student(id):
    if request.method == 'POST':
        name = request.form['name']
        grade = request.form['grade']
        card_id = request.form['card_id']

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""
            UPDATE students
            SET name = %s,
                grade = %s,
                card_id = %s
            WHERE id = %s
        """, (name, grade, card_id, id))
        flash('Student Updated Successfully')
        conn.commit()
        return redirect(url_for('adminpn.list_students'))

###########STUDENT DELETE##########


@adminpn.route('/adminpn/student_delete/<string:id>', methods=['POST', 'GET'])
def delete_student(id):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute('DELETE FROM students WHERE id = {0}'.format(id))
    conn.commit()
    flash('Student Removed Successfully')
    return redirect(url_for('adminpn.list_students'))
###################################

####################BOOKS STUFF##############################

#############BOOKS LIST#############


@adminpn.route('/adminpn/books', methods=['POST', 'GET'])
def books():
    if 'loggedin' in session:
        if session['role'] == 'Admin':

            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            s = "SELECT * FROM lib"
            cur.execute(s)  # Execute the SQL
            list_users = cur.fetchall()
        else:
            flash('You are not allowed')
            return redirect(url_for('userpn.dashboard'))

        return render_template('/admin/books.html', list_users=list_users, username=session['username'])
    return redirect(url_for('login'))
####################################

##############BOOK ADD##############


@adminpn.route('/adminpn/book_add', methods=['POST', 'GET'])
def add_book():

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        isbn = request.form['isbn']
        if len(isbn) < 13:
            flash("Must be 13 characters length")
            return redirect(url_for('adminpn.books'))
        else:
            try:
                h = {'Authorization': '47903_22f9fc8e767106bfc421df921b30c834'}
                z = requests.get(
                    f"https://api2.isbndb.com/book/{isbn}", headers=h)
                data = z.json()
                try:
                    json_expr_isbn = parse('$.*.isbn')
                    raw_isbn = json_expr_isbn.find(data)
                    isbn = raw_isbn[0].value
                except:
                    print("There is no isbn value")
                    isbn = ""
            # Finding and keeping the title
                try:
                    json_expr_title = parse('$.*.title')
                    raw_title = json_expr_title.find(data)
                    title = raw_title[0].value
                except:
                    print("There is no title value")
                    title = ""
            # Finding and keeping the publishers
                try:
                    json_expr_publishers = parse('$.*.publisher')
                    raw_publishers = json_expr_publishers.find(data)
                    publishers = raw_publishers[0].value
                except:
                    print("There is no publishers value")
                    publishers = ""
            # Finding and keeping the publish date
                try:
                    json_expr_publish_date = parse("$.*.date_published")
                    raw_publish_date = json_expr_publish_date.find(data)
                    publish_date = raw_publish_date[0].value
                except:
                    print("There is no publish date value")
                    publish_date = ""
            # Finding and keeping the description
                try:
                    json_expr_description = parse('$.*.overview')
                    raw_description = json_expr_description.find(data)
                    description = raw_description[0].value
                except:
                    print("There is no description value")
                    description = ""
            # Finding and keeping the number_of_pages
                try:
                    json_expr_number_of_pages = parse('$.*.pages')
                    raw_number_of_pages = json_expr_number_of_pages.find(data)
                    number_of_pages = raw_number_of_pages[0].value
                except:
                    print("There is no number_of_pages value")
                    number_of_pages = ""
                try:
                    json_expr_isbn13 = parse('$.*.isbn13')
                    raw_isbn13 = json_expr_isbn13.find(data)
                    isbn13 = raw_isbn13[0].value
                except:
                    print("There is no isbn13 value")
                    isbn13 = ""

                try:
                    json_expr_binding = parse('$.*.binding')
                    raw_binding = json_expr_binding.find(data)
                    binding = raw_binding[0].value
                except:
                    print("There is no binding value")
                    binding = ""

                try:
                    json_expr_language = parse('$.*.language')
                    raw_language = json_expr_language.find(data)
                    language = raw_language[0].value
                except:
                    print("There is no Language value")
                    language = ""
                try:
                    json_expr_dimensions = parse('$.*.dimensions')
                    raw_dimensions = json_expr_dimensions.find(data)
                    dimensions = raw_dimensions[0].value
                except:
                    print("There is no dimensions value")
                    dimensions = ""
                try:
                    json_expr_image = parse('$.*.image')
                    raw_image = json_expr_image.find(data)
                    image = raw_image[0].value
                    print(image)
                    urllib.request.urlretrieve(
                        f"{image}", f"/Users/superadmin/Desktop/apps/static/Photo/{isbn}.jpg")
                    image_url = f"Photo/{isbn13}.jpg"
                except:
                    print("There is no image value")
                    image_url= ""
            # Finding and keeping the authors
                try:
                    json_expr_authors = parse('$.*.authors[0]')
                    raw_authors = json_expr_authors.find(data)
                    authors = raw_authors[0].value

                except:
                    print("There is no authors value")
                    authors = ""
            # Disalowing some characters for database
                disallowed_characters = '''[{}:/],.-'/"();'''
                for character in disallowed_characters:
                    description = str(description)
                    description = (description.replace(character, ""))
                    publishers = str(publishers)
                    publishers = (publishers.replace(character, ""))
                    title = str(title)
                    title = (title.replace(character, ""))
                    authors = str(authors)
                    authors = (authors.replace(character, ""))                    
                if title=="" and publishers=="" and publish_date=="" and number_of_pages=="" and authors=="":
                    flash("No such book in the database")
                    return redirect(url_for('adminpn.books'))
                else:
                    cur.execute(
                        f"INSERT INTO lib (isbn , title , publishers , publish_date , description, number_of_pages, authors, isbn13, binding, language, dimensions, image)  VALUES  ('{isbn}' , '{title}' , '{publishers}' , '{publish_date}', '{description}', '{number_of_pages}', '{authors}', '{isbn13}', '{binding}' , '{language}', '{dimensions}', '{image_url}')")
                    conn.commit()
                 
                    flash('Book Added successfully')
                    return redirect(url_for('adminpn.books'))
            except:  
                flash("No internet or something gone wrong")
                return redirect(url_for('adminpn.books'))
@adminpn.route('/adminpn/book_add_man', methods=['POST', 'GET'])
def add_book_man_eng():
    return render_template("/admin/book_add_man_eng.html")


@adminpn.route('/adminpn/manual_submit', methods=['POST', 'GET'])
def manual_submit():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        isbn = request.form['isbn']
        name = request.form['name']
        author = request.form['author']
        quantity = request.form['quantity']
        language = request.form['language']
        cur.execute(
            f"INSERT INTO lib (isbn ,title,isbn13, authors,  language,quant)  VALUES  ('{isbn}','{name}' , '{isbn}','{author}', '{language}', '{quantity}')")
        conn.commit()
        flash('added')
        return redirect(url_for('adminpn.add_book_man_eng'))
    flash('no goodies')
    return redirect(url_for('adminpn.add_book_man_eng'))
####################################

##############BOOK DEL##############


@adminpn.route('/adminpn/deletebook/<string:id>', methods=['POST', 'GET'])
def delete_book(id):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("DELETE FROM lib WHERE isbn LIKE '{0}'".format(id))
    conn.commit()
    flash('Book Removed Successfully')
    return redirect(url_for('adminpn.books'))
####################################

####################ORDERS STUFF##############################

##############ORDER LIST##############


@adminpn.route('/adminpn/orders', methods=['POST', 'GET'])
def orders():
    if 'loggedin' in session:
        if session['role'] == 'Admin':
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            s = "SELECT * FROM orders"
            cur.execute(s) 
            list_users = cur.fetchall()
            return render_template("/admin/orders.html", list_users=list_users, username=session['username'])
        else:
            flash('You dont have an admin control, sorry dude')
            return redirect(url_for('userpn.dashboard'))

    return redirect(url_for('login'))
######################################


@adminpn.route('/adminpn/orders_staff', methods=['POST', 'GET'])
def orders_staff():
    if 'loggedin' in session:
        if session['role'] == 'Admin':
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            s = "SELECT * FROM orders_staff"

            cur.execute(s)  # Execute the SQL
            list_users = cur.fetchall()

            return render_template("/admin/orders_staff.html", list_users=list_users, username=session['username'])
        else:
            flash('You dont have an admin control, sorry dude')
            return redirect(url_for('userpn.dashboard'))

    return redirect(url_for('login'))
##############ADD ORDERS##############


@adminpn.route('/adminpn/add_orders', methods=['POST', 'GET'])
def add_orders():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if request.method == 'POST':
        oisbn = request.form['oisbn']
        oquantity = request.form['oquantity']
        oborrowed = request.form['oborrowed']
        oidCard = request.form['oidCard']
        if len(oisbn) == 13:
            try:
                datetime.datetime.strptime(oborrowed, '%d-%m-%Y')
                try:
                    cur.execute(
                        f"INSERT INTO orders (isbn, quantity, borrowed, card_id , name ,grade , photost) SELECT '{oisbn}','{oquantity}','{oborrowed}', '{oidCard}', name , grade, photo FROM students WHERE card_id= '{oidCard}'")
                    conn.commit()
                    telegram_send.send(
                        messages=[f"ID card = {oidCard} \n Borrowed = {oborrowed} \n ISBN = {oisbn}"])
                    flash('Order Added successfully')

                    return redirect(url_for('adminpn.orders'))

                except:
                    flash('No such student')
                    return redirect(url_for('adminpn.orders'))

            except:
                flash("Borrow date should be: DD-MM-YEAR")
                return redirect(url_for('adminpn.orders'))
        else:
            flash('ISBN is not right')
            return redirect(url_for('adminpn.orders'))
#################################

##############DELETE ORDER#######


@adminpn.route('/orderdel/<string:id>', methods=['POST', 'GET'])
def order_delete(id):
    id = int(id)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('DELETE FROM orders WHERE id = {0}'.format(id))
    conn.commit()
    flash('Order Removed Successfully')
    return redirect(url_for('adminpn.orders'))
##################################

################EDIT ORDER########


@adminpn.route('/adminpn/orderedit/<id>', methods=['POST', 'GET'])
def orderedit(id):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(f'SELECT * FROM orders WHERE id = {id}')
    data = cur.fetchall()
    cur.close()
    print(data[0])

    return render_template('/admin/order_edit.html', student=data[0])
#################################


##########UDATE ORDER############
@adminpn.route('/adminpn/orderupdate/<id>', methods=['POST', 'GET'])
def update_order(id):
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('adminpn.orders'))
    else:
        file = request.files['file']
    if file.filename == '':
        flash('No image selected for uploading')
        return redirect(url_for('adminpn.orders'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        photost = f'Photo/{filename}'
        returned = request.form['returned']
        condition = request.form['condition']
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(
            f"UPDATE orders SET returned = '{returned}', condition = '{condition}',afterlifeph = '{photost}' WHERE id = {id}")
        flash('Order Updated Successfully')
        conn.commit()
        if condition == 'Bad':
            telegram_send.send(
                messages=[f"Book is returned {returned} 😏 \n it is in {condition} condition 🤢"])
        elif condition == 'Good':
            telegram_send.send(
                messages=[f"Book is returned {returned} 😏 \n it is in {condition} condition 👍"])
        else:
            telegram_send.send(
                messages=[f"Book is returned {returned} 😏 \n it is in {condition} condition 👌"])
        return redirect(url_for('adminpn.orders'))
    else:
        flash('Allowed image types are - png, jpg, jpeg, gif')
        return redirect(url_for('adminpn.orders'))
#################################


@adminpn.route('/adminpn/teachers', methods=['POST', 'GET'])
def teachers():
    if 'loggedin' in session:
        if session['role'] == 'Admin':

            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            s = "SELECT * FROM teachers"
            cur.execute(s)  # Execute the SQL
            list_users = cur.fetchall()
            return render_template('/admin/teachers.html', list_users=list_users, username=session['username'])
        else:
            flash('You are not admin to see it!!!')
            return redirect(url_for('userpn.dashboard'))
    else:
        return redirect(url_for('login'))

#############################################################


@adminpn.route('/adminpn/logs', methods=['POST', 'GET'])
def logs():

    if 'loggedin' in session:
        if session['role'] == 'Admin':
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            return render_template('/admin/logs.html')
        else:
            flash('You are not allowed to see logs')
            return redirect(url_for('userpn.dashboard'))
    else:
        return redirect(url_for('login'))


@adminpn.route('/adminpn/log_for_students', methods=['POST', 'GET'])
def log_for_students():

    if 'loggedin' in session:
        if session['role'] == 'Admin':
            if os.path.exists("oreders_log.xlsx"):
                os.remove("orders_log.xlsx")
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute('SELECT COUNT(*) FROM orders')
            result = cur.fetchall()
            for row in result:
                number = row[0]
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Total orders'
            ws['A2'] = number
            wb.save("orders_log.xlsx")
            path = "orders_log.xlsx"
            return send_file(path, as_attachment=True)

        else:
            flash('You are not allowed to see logs')
            return redirect(url_for('userpn.dashboard'))
    else:
        return redirect(url_for('login'))
